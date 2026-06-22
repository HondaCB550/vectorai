"""
seed_equivalencias.py — Extrae matches confirmados de todos los Borradores
y los consolida en decisiones_usuario.json para la API VectorAI.

Uso:
    python seed_equivalencias.py [--dry-run]

Lee todos los .xlsx en C:\Pablo\Cotizaciones\Borradores\ que tengan una
hoja con columna DECISIÓN. Extrae filas CARGAR y CAMBIAR.
Consolida con el decisiones_usuario.json existente (sin duplicar).
Escribe el resultado en api/data/decisiones_usuario.json.
"""
import sys
import json
import argparse
from pathlib import Path

import openpyxl

BORRADORES_DIR = Path("C:/Pablo/Cotizaciones/Borradores")
OUTPUT_JSON    = Path(__file__).parent / "decisiones_usuario.json"

# Columnas esperadas en el borrador (fila de encabezado)
COL_N          = "N°"
COL_COD_PROV   = "COD PROV"
COL_DESC_PROV  = "DESC PROVEEDOR"
COL_CANT       = "CANT"
COL_PRECIO     = "PRECIO PDF"
COL_ORIGEN     = "ORIGEN"
COL_COD_INT    = "COD INTERNO PROP."
COL_SCORE      = "SCORE"
COL_DECISION   = "DECISIÓN"
COL_COD_CORRECTO = "CÓD. CORRECTO (si cambias)"


def _normalizar(s: str) -> str:
    return (s or "").upper().strip()


def _find_header_row(rows: list[tuple]) -> int | None:
    """Busca la fila que contiene COD PROV y DECISIÓN."""
    for i, row in enumerate(rows):
        vals = [str(v or "").upper().strip() for v in row]
        if any("COD PROV" in v for v in vals) and any("DECISI" in v for v in vals):
            return i
    return None


def _col_index(header: tuple, name: str) -> int | None:
    name_norm = name.upper().strip()
    for i, v in enumerate(header):
        if v and name_norm in str(v).upper().strip():
            return i
    return None


def extraer_de_borrador(path: Path) -> list[dict]:
    """Extrae entradas confirmadas de un archivo borrador."""
    resultados = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ⚠ No se pudo abrir {path.name}: {e}")
        return []

    for shname in wb.sheetnames:
        ws = wb[shname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        hi = _find_header_row(rows)
        if hi is None:
            continue

        header = rows[hi]
        ci = {
            "cod_prov":    _col_index(header, "COD PROV"),
            "desc_prov":   _col_index(header, "DESC PROVEEDOR"),
            "cant":        _col_index(header, "CANT"),
            "precio":      _col_index(header, "PRECIO PDF"),
            "origen":      _col_index(header, "ORIGEN"),
            "cod_int":     _col_index(header, "COD INTERNO PROP"),
            "score":       _col_index(header, "SCORE"),
            "decision":    _col_index(header, "DECISI"),
            "cod_correcto": _col_index(header, "CORRECTO"),
        }

        if ci["decision"] is None or ci["desc_prov"] is None:
            continue

        for row in rows[hi + 1:]:
            dec = str(row[ci["decision"]] or "").upper().strip() if row[ci["decision"]] else ""
            if dec not in ("CARGAR", "CAMBIAR"):
                continue

            desc = str(row[ci["desc_prov"]] or "").strip() if ci["desc_prov"] is not None else ""
            if not desc:
                continue

            cod_prov = str(row[ci["cod_prov"]] or "").strip() if ci["cod_prov"] is not None else ""
            cod_propuesto = str(row[ci["cod_int"]] or "").strip() if ci["cod_int"] is not None else ""
            cod_correcto = str(row[ci["cod_correcto"]] or "").strip() if ci["cod_correcto"] is not None else ""
            precio = row[ci["precio"]] if ci["precio"] is not None else 0
            cant = row[ci["cant"]] if ci["cant"] is not None else 1
            score = row[ci["score"]] if ci["score"] is not None else None
            origen = str(row[ci["origen"]] or "").strip() if ci["origen"] is not None else "IA"

            resultados.append({
                "cod_prov":      cod_prov,
                "desc_prov":     desc,
                "cant":          float(cant or 1),
                "precio_con_iva": float(precio or 0),
                "precio_sin_iva": round(float(precio or 0) / 1.105, 2),
                "cod_propuesto": cod_propuesto,
                "decision":      dec,
                "cod_correcto":  cod_correcto if cod_correcto else None,
                "score":         float(score) if score is not None else None,
                "origen":        origen,
                "fuente":        path.name,
            })

    wb.close()
    return resultados


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    # Cargar existentes (no duplicar)
    existing: list[dict] = []
    if OUTPUT_JSON.exists():
        with open(OUTPUT_JSON, encoding="utf-8") as f:
            existing = json.load(f)

    # Clave de deduplicación: desc_prov normalizada + cod efectivo
    def _key(e: dict) -> str:
        cod = (e.get("cod_correcto") or e.get("cod_propuesto") or "").strip()
        desc = (e.get("desc_prov") or "").upper().strip()
        return f"{desc}|{cod}"

    existing_keys = {_key(e) for e in existing}
    print(f"Existentes: {len(existing)} entradas")

    # Leer todos los borradores
    all_new = []
    xlsx_files = sorted(BORRADORES_DIR.glob("*.xlsx"))
    print(f"Procesando {len(xlsx_files)} archivos en {BORRADORES_DIR}...\n")

    for path in xlsx_files:
        entradas = extraer_de_borrador(path)
        nuevas = [e for e in entradas if _key(e) not in existing_keys]
        if entradas:
            print(f"  {path.name}: {len(entradas)} CARGAR/CAMBIAR -> {len(nuevas)} nuevas")
        all_new.extend(nuevas)
        for e in nuevas:
            existing_keys.add(_key(e))

    print(f"\nTotal nuevas: {len(all_new)}")

    if not all_new:
        print("Nada nuevo para agregar.")
        return

    # Renumerar y combinar
    combined = existing + all_new
    for i, e in enumerate(combined, start=1):
        e["n"] = i

    print(f"Total combinado: {len(combined)} entradas")

    # Breakdown por categoría de cod_int
    from collections import Counter
    efectivos = [e.get("cod_correcto") or e.get("cod_propuesto") or "" for e in all_new
                 if e.get("decision") == "CARGAR"]
    prefijos = Counter(c[:4] for c in efectivos if c)
    print("\nNuevas CARGAR por prefijo de código:")
    for p, n in prefijos.most_common(20):
        print(f"  {p}: {n}")

    if args.dry_run:
        print("\n[DRY RUN] No se escribió nada.")
        return

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(combined, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Escrito: {OUTPUT_JSON}")
    print(f"  {len(combined)} entradas totales ({len(all_new)} nuevas)")


if __name__ == "__main__":
    main()
