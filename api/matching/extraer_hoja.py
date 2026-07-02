"""extraer_hoja.py — extracción de ítems desde planillas (.xlsx, .csv, Google Sheets).

Import de Google Sheets sin API key: se usa la URL de export del documento
(requiere que el link esté compartido como "Cualquiera con el enlace").

Devuelve el mismo dict que extraer_pdf_texto.extraer().
"""
import csv
import io
import re
import unicodedata

from extraer_pdf_texto import parse_num

# Aliases de encabezado → campo interno
HDR = {
    "cod": {"cod", "codigo", "cod prov", "ref", "referencia", "item", "articulo", "art", "sku"},
    "desc": {"descripcion", "detalle", "producto", "material", "denominacion", "descripcion del articulo", "concepto"},
    "cant": {"cant", "cantidad", "cant.", "unidades", "un", "q"},
    "pu": {"precio", "precio unit", "precio unitario", "p unit", "p. unit", "unitario", "precio unit.", "p.unitario", "pu"},
    "total": {"total", "importe", "subtotal", "imp total", "total unif", "importe total"},
}

RE_GSHEET_ID = re.compile(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)")
RE_GID = re.compile(r"[#?&]gid=(\d+)")


def _norm(s) -> str:
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.replace("$", "").replace(":", "")).strip()


def _mapear_header(fila: list) -> dict | None:
    """Devuelve {campo: índice_columna} si la fila parece un encabezado."""
    mapa = {}
    for i, celda in enumerate(fila):
        n = _norm(celda)
        if not n:
            continue
        for campo, aliases in HDR.items():
            if campo not in mapa and (n in aliases or any(n.startswith(a + " ") for a in aliases)):
                mapa[campo] = i
                break
    # Un encabezado real tiene al menos descripción + algún precio
    if "desc" in mapa and ("pu" in mapa or "total" in mapa):
        return mapa
    return None


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return parse_num(str(v))
    except (ValueError, TypeError):
        return 0.0


def _extraer_filas(filas: list[list]) -> dict:
    """Núcleo compartido: detecta encabezado y extrae ítems de una matriz."""
    mapa, inicio = None, 0
    for i, fila in enumerate(filas[:30]):
        mapa = _mapear_header(fila)
        if mapa:
            inicio = i + 1
            break
    if not mapa:
        raise ValueError(
            "No encontré el encabezado de la tabla (busco columnas tipo "
            "DESCRIPCIÓN + PRECIO/TOTAL en las primeras 30 filas)."
        )

    def celda(fila, campo):
        idx = mapa.get(campo)
        return fila[idx] if idx is not None and idx < len(fila) else None

    items = []
    for fila in filas[inicio:]:
        desc = str(celda(fila, "desc") or "").strip()
        if not desc or _norm(desc) in ("total", "subtotal", "iva", "total general"):
            continue
        pu = _num(celda(fila, "pu"))
        total = _num(celda(fila, "total"))
        cant = _num(celda(fila, "cant")) or 1.0
        if pu <= 0 and total > 0:
            pu = round(total / cant, 2)
        if total <= 0 and pu > 0:
            total = round(pu * cant, 2)
        if pu <= 0:
            continue
        items.append({
            "cod": str(celda(fila, "cod") or "").strip(),
            "desc": desc, "cant": cant, "pu": pu, "total": total,
        })

    return {
        "fecha_presupuesto": None,
        "iva_detectado": "ASUMIDO 1,105",
        "suma_lineas": round(sum(it["total"] for it in items), 2),
        "n_items": len(items),
        "metodo_extraccion": "planilla",
        "items": items,
    }


def extraer_xlsx(content: bytes) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        filas = [list(r) for r in ws.iter_rows(values_only=True, max_row=2000)]
    finally:
        wb.close()
    return _extraer_filas(filas)


def extraer_csv(content: bytes) -> dict:
    texto = content.decode("utf-8-sig", errors="replace")
    # Detectar separador (coma vs punto y coma, común en CSVs argentinos)
    try:
        dialect = csv.Sniffer().sniff(texto[:2000], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    filas = list(csv.reader(io.StringIO(texto), dialect))
    return _extraer_filas(filas)


def descargar_gsheet(url: str) -> tuple[str, bytes]:
    """Descarga un Google Sheet compartido por link como xlsx.
    Devuelve (nombre_para_mostrar, bytes_xlsx)."""
    import requests

    m = RE_GSHEET_ID.search(url or "")
    if not m:
        raise ValueError("El link no parece de Google Sheets (falta /spreadsheets/d/...).")
    sheet_id = m.group(1)
    export = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    g = RE_GID.search(url)
    if g:
        export += f"&gid={g.group(1)}"

    resp = requests.get(export, timeout=30, allow_redirects=True)
    ctype = resp.headers.get("content-type", "")
    if resp.status_code in (401, 403) or "text/html" in ctype:
        raise ValueError(
            "No pude acceder a la planilla. En Google Sheets: Compartir → "
            "'Cualquier persona con el enlace' (como lector) y volvé a intentar."
        )
    resp.raise_for_status()
    return f"GoogleSheet_{sheet_id[:8]}.xlsx", resp.content
