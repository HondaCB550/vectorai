"""
exportar_excel.py — Genera un Excel comparativo de presupuestos Vectorai
Devuelve bytes del .xlsx, listo para enviar como respuesta HTTP.
Encabezado de marca (isologo + wordmark embebido como imagen) desde marca.py.
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import marca
from lista_compras import pedidos_por_proveedor, subtotal_fila

# Paleta
COLOR_NAVY     = marca.NAVY_HEX
COLOR_GRIS     = marca.GRIS_HEX
COLOR_SUBHDR   = "EDF0F5"   # gris claro
COLOR_MEJOR    = "C6F6D5"   # verde claro
COLOR_RUBRO    = "E8EDF5"   # azul muy claro
COLOR_BLANCO   = "FFFFFF"
COLOR_TEXTO_W  = "FFFFFF"

COLORES_PROV = [
    "3373C2",  # azul
    "D45C00",  # naranja
    "2EA15E",  # verde
    "993399",  # violeta
    "BF2626",  # rojo
    "4D99B3",  # celeste
]

NUM_FMT = '"$" #,##0'


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)


def _border_bottom() -> Border:
    side = Side(style="medium", color="B0B5BE")
    return Border(bottom=side)


# Lockup del isologo renderizado una sola vez (a 3x para que quede nítido)
_LOCKUP_PNG, _LOCKUP_W, _LOCKUP_H = marca.lockup_png(alto=72)
_LOCKUP_ALTO_FINAL = 24  # px en la hoja


def _logo_imagen() -> XLImage:
    """Instancia nueva del lockup por hoja (openpyxl no comparte imágenes)."""
    img = XLImage(BytesIO(_LOCKUP_PNG))
    img.height = _LOCKUP_ALTO_FINAL
    img.width  = round(_LOCKUP_W * _LOCKUP_ALTO_FINAL / _LOCKUP_H)
    return img


def _encabezado_marca(ws, total_cols: int, titulo: str, subtitulo: str, ancla: str = "B1"):
    """Filas 1-3: logo, título navy y metadatos con regla naranja. Fondo blanco."""
    ws.add_image(_logo_imagen(), ancla)
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c = ws.cell(2, 1, titulo)
    c.font = _font(bold=True, color=COLOR_NAVY, size=13)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=2)
    ws.row_dimensions[2].height = 22

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    c = ws.cell(3, 1, subtitulo)
    c.font = _font(color=COLOR_GRIS, size=9)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=2)
    borde_naranja = Border(bottom=Side(style="medium", color=marca.NARANJA_HEX))
    for col in range(1, total_cols + 1):
        ws.cell(3, col).border = borde_naranja
    ws.row_dimensions[3].height = 16


def generar_excel_comparativo(
    comparativo: list[dict],
    proveedores: list[str],
    titulo: str = None,
    subtitulo: str = None,
    config_proveedores: dict = None,
    vista_efectivo: bool = False,
) -> bytes:
    """
    Genera un Excel comparativo y devuelve los bytes del archivo.

    En vista_efectivo (regla de Pablo, 10-07): el proveedor que cotizó c/IVA
    lleva UNA sola columna (su precio final — no puede vender sin IVA, mostrar
    un neto sería ficción); el que cotizó s/IVA lleva DOS columnas (s/IVA, que
    es la comparable, + c/IVA de referencia). El "mejor" se marca sobre las
    columnas comparables.
    """
    titulo = titulo or f"Vectorai — Comparativa {datetime.now().strftime('%Y-%m-%d')}"
    fecha  = datetime.now().strftime("%d/%m/%Y")
    subtitulo = subtitulo or f"Generado el {fecha} · Precios sin IVA · Vectorai"

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa"

    # Spec de columnas de proveedor: (proveedor, tipo)
    #   "principal" = el precio de la vista (comparable, recibe el "mejor")
    #   "ref"       = referencia c/IVA del que cotizó s/IVA (solo vista efectivo)
    cfg = config_proveedores or {}

    def _prov_con_iva(p: str) -> bool:
        return bool((cfg.get(p) or {}).get("iva_incluido", True))

    cols_prov: list[tuple[str, str, str]] = []   # (prov, tipo, etiqueta)
    if vista_efectivo and cfg:
        for p in proveedores:
            if _prov_con_iva(p):
                cols_prov.append((p, "principal", f"{p} (c/IVA)"))
            else:
                cols_prov.append((p, "principal", f"{p} (s/IVA)"))
                cols_prov.append((p, "ref", f"{p} (c/IVA ref.)"))
    else:
        cols_prov = [(p, "principal", p) for p in proveedores]

    # Columnas: Rubro(oculto) | Material | Cant. | Unidad | provs... | Mejor precio | Ahorro
    n_cols_prov = len(cols_prov)
    col_rubro  = 1
    col_mat    = 2
    col_cant   = 3
    col_unidad = 4
    col_prov0  = 5
    col_mejor  = col_prov0 + n_cols_prov
    col_ahorro = col_mejor + 1
    total_cols = col_ahorro

    # ── Filas 1-3: encabezado de marca (logo + título + metadatos) ────────────
    _encabezado_marca(ws, total_cols,
                      marca.titulo_visible(titulo),
                      marca.meta_visible(subtitulo))

    # ── Fila 4: encabezados de columnas ───────────────────────────────────────
    etiqueta_ahorro = "Ahorro" if (vista_efectivo and cfg) else "Ahorro s/IVA"
    headers = ["Rubro", "Material", "Cant.", "Unidad"] + [e for _, _, e in cols_prov] \
        + ["Mejor precio", etiqueta_ahorro]
    # Color por PROVEEDOR (las dos columnas del mismo proveedor comparten color)
    color_prov = {p: COLORES_PROV[i % len(COLORES_PROV)] for i, p in enumerate(proveedores)}
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(4, col_idx, label)
        c.alignment = Alignment(horizontal="center" if col_idx > 3 else "left",
                                vertical="center", indent=1 if col_idx <= 3 else 0)
        c.border = _border_bottom()
        if col_idx in (col_rubro, col_mat, col_cant, col_unidad):
            c.fill = _fill(COLOR_SUBHDR)
            c.font = _font(bold=True, size=9)
        elif col_idx < col_mejor:
            prov, tipo, _ = cols_prov[col_idx - col_prov0]
            c.fill = _fill(color_prov[prov])
            c.font = _font(bold=(tipo == "principal"), color=COLOR_TEXTO_W, size=9,
                           italic=(tipo == "ref"))
        else:
            c.fill = _fill(COLOR_SUBHDR)
            c.font = _font(bold=True, size=9)
    ws.row_dimensions[4].height = 20

    # Freeze rows 1-4 y columna Material
    ws.freeze_panes = "C5"

    # ── Filas de datos ────────────────────────────────────────────────────────
    current_row = 5
    ultimo_rubro = None
    totales = [0.0] * n_cols_prov   # un total por columna (incluye las de referencia)

    for row in comparativo:
        rubro   = row.get("rubro", "")
        mat     = row.get("material", "")
        cant    = row.get("cant", 1) or 1
        unidad  = row.get("unidad", "")
        precios = row.get("precios", {})
        mejor   = row.get("mejor_proveedor", "")
        ahorro  = row.get("ahorro", 0) or 0

        precios_vals = {p: precios[p]["precio_sin_iva"] * cant for p in proveedores if p in precios}
        precio_min = min(precios_vals.values()) if len(precios_vals) > 1 else None

        # Separador de rubro
        if rubro != ultimo_rubro:
            if ultimo_rubro is not None:
                current_row += 1  # fila vacía entre rubros
            for col in range(1, total_cols + 1):
                c = ws.cell(current_row, col)
                c.fill = _fill(COLOR_RUBRO)
                if col == col_mat:
                    c.value = rubro.upper()
                    c.font  = _font(bold=True, size=9, color="3A5080")
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[current_row].height = 16
            current_row += 1
            ultimo_rubro = rubro

        # Fila de datos
        ws.cell(current_row, col_rubro, "").fill = _fill(COLOR_BLANCO)

        c = ws.cell(current_row, col_mat, mat)
        c.font = _font(size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)

        c = ws.cell(current_row, col_cant, cant if cant != 1 else "")
        c.font = _font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(current_row, col_unidad, unidad)
        c.font = _font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")

        for i, (prov, tipo, _) in enumerate(cols_prov):
            col = col_prov0 + i
            if tipo == "ref":
                pu_ref = precios.get(prov, {}).get("precio_con_iva")
                val = round(pu_ref * cant, 2) if pu_ref is not None else None
            else:
                val = precios_vals.get(prov)
            is_mejor = (tipo == "principal" and precio_min is not None
                        and val is not None and val == precio_min
                        and len(precios_vals) > 1)
            if val is not None:
                totales[i] += val
                c = ws.cell(current_row, col, val)
                c.number_format = NUM_FMT
            else:
                c = ws.cell(current_row, col, "—")
            c.font = _font(bold=is_mejor, size=9, italic=(tipo == "ref"),
                           color="808080" if tipo == "ref" else "000000")
            c.alignment = Alignment(horizontal="right", vertical="center")
            if is_mejor:
                c.fill = _fill(COLOR_MEJOR)

        # Mejor proveedor
        c = ws.cell(current_row, col_mejor, mejor or "")
        c.font = _font(bold=bool(mejor), size=9,
                       color="1A804A" if mejor else "000000")
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Ahorro
        if ahorro:
            c = ws.cell(current_row, col_ahorro, ahorro)
            c.number_format = NUM_FMT
            c.font = _font(size=9, color="1A804A")
        else:
            ws.cell(current_row, col_ahorro, "")
        ws.cell(current_row, col_ahorro).alignment = Alignment(
            horizontal="right", vertical="center")

        ws.row_dimensions[current_row].height = 15
        current_row += 1

    # ── Fila de totales ───────────────────────────────────────────────────────
    current_row += 1
    ws.cell(current_row, col_mat, "TOTAL (matches OK + REVISAR)").font = _font(bold=True, size=9)
    ws.cell(current_row, col_mat).fill = _fill(COLOR_SUBHDR)
    ws.cell(current_row, col_mat).alignment = Alignment(horizontal="left", indent=2)
    ws.cell(current_row, col_rubro).fill = _fill(COLOR_SUBHDR)
    ws.cell(current_row, col_unidad).fill = _fill(COLOR_SUBHDR)

    for i, (prov, tipo, _) in enumerate(cols_prov):
        col = col_prov0 + i
        c = ws.cell(current_row, col, round(totales[i], 2))
        c.number_format = NUM_FMT
        c.font = _font(bold=(tipo == "principal"), size=9, italic=(tipo == "ref"))
        c.fill = _fill(COLOR_SUBHDR)
        c.alignment = Alignment(horizontal="right")

    ws.cell(current_row, col_mejor).fill = _fill(COLOR_SUBHDR)
    ws.cell(current_row, col_ahorro).fill = _fill(COLOR_SUBHDR)
    ws.row_dimensions[current_row].height = 18

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(col_rubro)].width  = 0.5
    ws.column_dimensions[get_column_letter(col_mat)].width    = 42
    ws.column_dimensions[get_column_letter(col_cant)].width   = 7
    ws.column_dimensions[get_column_letter(col_unidad)].width = 9
    for i in range(n_cols_prov):
        ws.column_dimensions[get_column_letter(col_prov0 + i)].width = 18
    ws.column_dimensions[get_column_letter(col_mejor)].width  = 18
    ws.column_dimensions[get_column_letter(col_ahorro)].width = 16

    # ── Hojas de compras: una por proveedor con sus ítems ganadores ──────────
    _agregar_hojas_compras(wb, comparativo, proveedores, subtitulo)

    # ── Exportar a bytes ──────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


_INVALIDOS_HOJA = set('[]:*?/\\')


def _nombre_hoja(nombre: str) -> str:
    limpio = "".join(ch for ch in (nombre or "") if ch not in _INVALIDOS_HOJA).strip()
    return (f"Pedido {limpio}")[:31] or "Pedido"


def _agregar_hojas_compras(wb, comparativo: list[dict], proveedores: list[str], subtitulo: str = None):
    """Listado de compras: una hoja por proveedor con SOLO los materiales donde
    ese proveedor tiene el mejor precio — el pedido listo para mandarle.

    El agrupado sale de lista_compras.pedidos_por_proveedor(), compartido con
    el PDF y la imagen: los tres archivos arman el mismo pedido. Las hojas
    quedan ordenadas por total descendente (igual que el tab en pantalla), no
    por el orden en que se subieron los proveedores.
    """
    for pedido in pedidos_por_proveedor(comparativo):
        prov  = pedido["proveedor"]
        filas = pedido["filas"]
        ws = wb.create_sheet(_nombre_hoja(prov))

        meta = ((marca.meta_visible(subtitulo) + " · ") if subtitulo else "") + \
               "Solo ítems donde este proveedor tiene el mejor precio"
        _encabezado_marca(ws, 5, f"Pedido — {prov}", meta, ancla="A1")

        headers = ["Material", "Cant.", "Unidad", "Precio unit.", "Subtotal"]
        for j, h in enumerate(headers, start=1):
            c = ws.cell(4, j, h)
            c.font = _font(bold=True, size=9)
            c.fill = _fill(COLOR_SUBHDR)
            c.alignment = Alignment(horizontal="left" if j == 1 else "right")

        fila = 5
        total = pedido["total"]
        for r in filas:
            precio, cant, subtotal = subtotal_fila(r, prov)
            ws.cell(fila, 1, r.get("material", "")).font = _font(size=9)
            for j, (val, fmt) in enumerate([(cant, None), (r.get("unidad", ""), None),
                                            (precio, NUM_FMT), (subtotal, NUM_FMT)], start=2):
                c = ws.cell(fila, j, val)
                c.font = _font(size=9)
                c.alignment = Alignment(horizontal="right")
                if fmt:
                    c.number_format = fmt
            fila += 1

        fila += 1
        ws.cell(fila, 1, f"TOTAL PEDIDO ({len(filas)} ítems)").font = _font(bold=True, size=10)
        for j in range(1, 5):
            ws.cell(fila, j).fill = _fill(COLOR_MEJOR)
        c = ws.cell(fila, 5, round(total, 2))
        c.number_format = NUM_FMT
        c.font = _font(bold=True, size=10)
        c.fill = _fill(COLOR_MEJOR)
        c.alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 46
        for col, w in (("B", 8), ("C", 9), ("D", 14), ("E", 14)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A5"
