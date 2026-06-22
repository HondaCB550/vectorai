"""
Migración: extrae todos los matches de los borradores revisados → material_denominaciones.
Lee DESC PROVEEDOR + COD INT PROP de cada borrador y hace upsert en Supabase.
"""
import os, sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()
BORRADORES_DIR = r"C:\Pablo\Cotizaciones\Borradores"
SUPABASE_URL   = os.getenv("SUPABASE_URL")
SUPABASE_KEY   = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

# Codigos validos
res = sb.table("materiales_validados").select("codigo").execute()
codigos_validos = {r["codigo"] for r in (res.data or [])}
print(f"{len(codigos_validos)} codigos validos en materiales_validados")

xlsx_files = sorted(f for f in os.listdir(BORRADORES_DIR) if f.endswith(".xlsx"))
print(f"Procesando {len(xlsx_files)} borradores...\n")

batch = []
seen  = set()
sin_cod = 0

for filename in xlsx_files:
    path = os.path.join(BORRADORES_DIR, filename)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR {filename}: {e}")
        continue

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ("como usar", "alternativas", "hoja1", "sheet1", "comparativa"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Encontrar header
        header_idx = None
        for i, row in enumerate(rows[:15]):
            row_str = " ".join(str(c or "") for c in row)
            if "COD INT" in row_str or "DESC PROVEEDOR" in row_str:
                header_idx = i
                break
        if header_idx is None:
            continue

        header = [str(c or "").strip().upper() for c in rows[header_idx]]

        try:
            idx_desc = next(i for i, h in enumerate(header) if "DESC" in h and "PROV" in h)
        except StopIteration:
            continue

        idx_cod = None
        for label in ("COD INT PROP", "COD INT", "CODIGO INTERNO", "COD INTERNO"):
            try:
                idx_cod = next(i for i, h in enumerate(header) if label in h)
                break
            except StopIteration:
                continue
        if idx_cod is None:
            continue

        # Proveedor del encabezado
        proveedor = sheet_name.upper().strip()
        for r in rows[:6]:
            if r and str(r[0] or "").upper() == "PROVEEDOR" and r[1]:
                proveedor = str(r[1]).upper().strip()
                break

        origen = f"borrador_{proveedor.lower().replace(' ', '_')}"
        filas_ok = 0

        for row in rows[header_idx+1:]:
            if not row or all(c is None for c in row):
                continue

            desc = str(row[idx_desc] or "").strip() if idx_desc < len(row) else ""
            cod  = str(row[idx_cod]  or "").strip() if idx_cod  < len(row) else ""

            if not desc or not cod or cod not in codigos_validos:
                if desc and not cod:
                    sin_cod += 1
                continue

            desc_norm = desc.lower().strip()
            key = (cod, desc_norm)
            if key in seen:
                continue
            seen.add(key)

            batch.append({
                "codigo_material":       cod,
                "denominacion":          desc_norm,
                "origen":                origen,
                "confianza":             90,
                "frecuencia_encontrada": 1,
            })
            filas_ok += 1

        if filas_ok:
            print(f"  {filename[:48]:48s}  +{filas_ok}")

print(f"\nTotal aliases a insertar (deduplicados): {len(batch)}")
print(f"Sin codigo interno (SIN MATCH/REVISAR):   {sin_cod}\n")

# Insertar en lotes
BATCH = 100
insertados = 0
for i in range(0, len(batch), BATCH):
    chunk = batch[i:i+BATCH]
    try:
        sb.table("material_denominaciones").upsert(
            chunk,
            on_conflict="codigo_material,denominacion"
        ).execute()
        insertados += len(chunk)
        pct = round(100 * (i + len(chunk)) / len(batch))
        print(f"  {i+len(chunk):4d}/{len(batch)}  ({pct}%)")
    except Exception as e:
        print(f"  ERROR lote {i}: {e}")

print(f"\nOK: {insertados} aliases insertados/actualizados")
